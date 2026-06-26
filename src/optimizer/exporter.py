import io
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

DAY_NAMES = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]

# Renkler
HEADER_COLOR = "1F4E79"
SUBHEADER_COLOR = "2E75B6"
ALT_ROW_COLOR = "D6E4F0"
WHITE = "FFFFFF"
EXAM_HEADER_COLOR = "7B3FA0"
EXAM_ALT_COLOR = "E8D5F5"

def make_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)

def style_header(cell, bg_color=HEADER_COLOR, font_size=11):
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.font = Font(bold=True, color=WHITE, size=font_size)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = make_border()

def style_cell(cell, bg_color=WHITE, bold=False, center=False):
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center", wrap_text=True
    )
    cell.border = make_border()

def generate_excel(schedule, exam_schedule):
    wb = Workbook()

    # ─── SAYFA 1: HAFTALIK DERS PROGRAMI ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Ders Programı"

    # Başlık
    ws1.merge_cells("A1:H1")
    title_cell = ws1["A1"]
    title_cell.value = "HAFTALIK DERS PROGRAMI"
    style_header(title_cell, HEADER_COLOR, 14)
    ws1.row_dimensions[1].height = 35

    # Sütun başlıkları
    headers = ["Kod", "Ders Adı", "Gün", "Başlangıç", "Bitiş", "Süre", "Derslik", "Öğretim Üyesi"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=2, column=col, value=header)
        style_header(cell, SUBHEADER_COLOR, 10)
    ws1.row_dimensions[2].height = 25

    # Veriler
    for row_idx, entry in enumerate(schedule, 3):
        bg = ALT_ROW_COLOR if row_idx % 2 == 0 else WHITE
        values = [
            entry.get("code", ""),
            entry.get("name", ""),
            DAY_NAMES[entry.get("dayOfWeek", 0)],
            f"{entry.get('startHour', '')}:00",
            f"{entry.get('endHour', '')}:00",
            f"{entry.get('durationHours', '')} saat",
            entry.get("classroom", ""),
            entry.get("instructor", ""),
        ]
        for col, value in enumerate(values, 1):
            cell = ws1.cell(row=row_idx, column=col, value=value)
            style_cell(cell, bg, bold=(col == 1), center=(col in [1, 3, 4, 5, 6]))
        ws1.row_dimensions[row_idx].height = 20

    # Sütun genişlikleri
    col_widths = [12, 30, 15, 12, 12, 12, 12, 25]
    for col, width in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(col)].width = width

    # ─── SAYFA 2: TAKVİM GÖRÜNÜMÜ ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Takvim Görünümü")
    hours = list(range(8, 19))

    # Başlık
    ws2.merge_cells(f"A1:{get_column_letter(len(DAY_NAMES) + 1)}1")
    t2 = ws2["A1"]
    t2.value = "HAFTALIK TAKVİM"
    style_header(t2, HEADER_COLOR, 14)
    ws2.row_dimensions[1].height = 35

    # Gün başlıkları
    ws2.cell(row=2, column=1, value="Saat")
    style_header(ws2.cell(row=2, column=1), SUBHEADER_COLOR)
    for col, day in enumerate(DAY_NAMES, 2):
        cell = ws2.cell(row=2, column=col, value=day)
        style_header(cell, SUBHEADER_COLOR)
    ws2.row_dimensions[2].height = 25

    # Saat satırları
    for row_idx, hour in enumerate(hours, 3):
        cell = ws2.cell(row=row_idx, column=1, value=f"{hour}:00")
        style_cell(cell, ALT_ROW_COLOR, bold=True, center=True)
        ws2.row_dimensions[row_idx].height = 30

        for col, day_idx in enumerate(range(5), 2):
            entry = next(
                (e for e in schedule if e.get("dayOfWeek") == day_idx and e.get("startHour") == hour),
                None
            )
            cell = ws2.cell(row=row_idx, column=col)
            if entry:
                cell.value = f"{entry.get('code')}\n{entry.get('classroom')}"
                style_cell(cell, "D6E4F0", bold=True, center=True)
            else:
                style_cell(cell, WHITE, center=True)

    # Sütun genişlikleri
    ws2.column_dimensions["A"].width = 10
    for col in range(2, 7):
        ws2.column_dimensions[get_column_letter(col)].width = 20

    # ─── SAYFA 3: SINAV TAKVİMİ ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Sınav Takvimi")

    ws3.merge_cells("A1:G1")
    t3 = ws3["A1"]
    t3.value = "SINAV TAKVİMİ"
    style_header(t3, EXAM_HEADER_COLOR, 14)
    ws3.row_dimensions[1].height = 35

    exam_headers = ["Sınav Günü", "Başlangıç", "Bitiş", "Kod", "Ders Adı", "Derslik", "Öğretim Üyesi"]
    for col, header in enumerate(exam_headers, 1):
        cell = ws3.cell(row=2, column=col, value=header)
        style_header(cell, "5B2C8D", 10)
    ws3.row_dimensions[2].height = 25

    for row_idx, exam in enumerate(exam_schedule, 3):
        bg = EXAM_ALT_COLOR if row_idx % 2 == 0 else WHITE
        values = [
            f"{exam.get('examDay')}. Gün",
            f"{exam.get('startHour')}:00",
            f"{exam.get('endHour')}:00",
            exam.get("code", ""),
            exam.get("name", ""),
            exam.get("classroom", ""),
            exam.get("instructor", ""),
        ]
        for col, value in enumerate(values, 1):
            cell = ws3.cell(row=row_idx, column=col, value=value)
            style_cell(cell, bg, bold=(col == 4), center=(col in [1, 2, 3, 4]))
        ws3.row_dimensions[row_idx].height = 20

    exam_col_widths = [12, 12, 12, 12, 30, 12, 25]
    for col, width in enumerate(exam_col_widths, 1):
        ws3.column_dimensions[get_column_letter(col)].width = width

    # Dosyayı byte olarak döndür
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()